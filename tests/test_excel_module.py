"""SpringBootAI Excel 模块测试 —— 覆盖注解/转换器/读写引擎/多 sheet/大数字/降级/round-trip。

使用真实 openpyxl 读写临时 xlsx 文件做 round-trip 验证，不依赖网络。
对齐 tests/test_ai_module.py 的 pytest 风格（class TestXxx + test_* 方法）。
"""
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest

PROJECT_ROOT = str(Path(__file__).parent.parent)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from spring.excel import (
    EasyExcel, ExcelProperty, ExcelIgnore, excel_sheet, ExcelSheet,
    read_excel, write_excel,
    Converter, StringConverter, IntegerConverter, FloatConverter,
    BooleanConverter, DateStringConverter, BigDecimalConverter, resolve_converter,
    ExcelPropertyError, ExcelReadError, ExcelWriteError, ExcelDependencyError,
)
from spring.excel.annotations import (
    parse_excel_columns, _get_class_sheet_meta, _field_to_header,
)


# ==================== 测试用实体 ====================

@ExcelSheet("用户列表", head_row_number=1)
class DemoUser:
    id = ExcelProperty("ID", order=1, big_number=True)
    name = ExcelProperty("姓名", order=2, width=12)
    age = ExcelProperty("年龄", order=3)
    amount = ExcelProperty("金额", order=4, converter=BigDecimalConverter)
    created_at = ExcelProperty("创建时间", order=5, date_format="%Y-%m-%d %H:%M:%S")
    active = ExcelProperty("是否启用", order=6)
    remark = ExcelIgnore()

    def __init__(self, id=None, name=None, age=None, amount=None,
                 created_at=None, active=None, remark=None):
        self.id = id
        self.name = name
        self.age = age
        self.amount = amount
        self.created_at = created_at
        self.active = active
        self.remark = remark

    def __repr__(self):
        return f"DemoUser(id={self.id!r}, name={self.name!r}, age={self.age!r})"


class PlainUser:
    """无任何注解的纯 __init__ 模型（回退路径）。"""
    def __init__(self, id=None, username=None, email=None):
        self.id = id
        self.username = username
        self.email = email


# ==================== 注解与元数据解析 ====================

class TestAnnotations:
    def test_excel_property_metadata(self):
        p = ExcelProperty("姓名", order=2, width=15, big_number=True)
        assert p.value == "姓名"
        assert p.order == 2
        assert p.width == 15
        assert p.big_number is True

    def test_excel_property_header_resolution(self):
        p = ExcelProperty()
        assert p.resolve_header("user_name") == "User Name"
        assert p.resolve_header("userName") == "User Name"
        assert ExcelProperty("自定义").resolve_header("x") == "自定义"

    def test_field_to_header(self):
        assert _field_to_header("id") == "Id"
        assert _field_to_header("created_at") == "Created At"

    def test_excel_ignore_marker(self):
        ign = ExcelIgnore()
        assert ign.attr_name == ""

    def test_excel_sheet_class_meta(self):
        meta = _get_class_sheet_meta(DemoUser)
        assert isinstance(meta, ExcelSheet)
        assert meta.sheet_name == "用户列表"
        assert meta.head_row_number == 1
        assert meta.freeze_head is True

    def test_excel_sheet_default_meta(self):
        meta = _get_class_sheet_meta(PlainUser)
        assert meta.sheet_name == ""
        assert meta.head_row_number == 1

    def test_parse_columns_order_and_ignore(self):
        cols = parse_excel_columns(DemoUser)
        names = [c.attr_name for c in cols]
        assert "remark" not in names  # 被 @ExcelIgnore 跳过
        assert names == ["id", "name", "age", "amount", "created_at", "active"]
        assert [c.header for c in cols] == ["ID", "姓名", "年龄", "金额", "创建时间", "是否启用"]

    def test_parse_columns_index_overrides_order(self):
        class C:
            a = ExcelProperty("A", order=1)
            b = ExcelProperty("B", order=2, index=0)
            def __init__(self, a=None, b=None):
                self.a = a; self.b = b
        cols = parse_excel_columns(C)
        # index=0 的 b 排在最前
        assert cols[0].attr_name == "b"
        assert cols[1].attr_name == "a"

    def test_parse_columns_fallback_no_annotation(self):
        cols = parse_excel_columns(PlainUser)
        assert [c.attr_name for c in cols] == ["id", "username", "email"]
        assert cols[0].header == "Id"

    def test_function_decorator_form(self):
        class Func:
            @ExcelProperty("姓名", order=1)
            def name(self): ...
            @ExcelProperty("年龄", order=2)
            def age(self): ...
            @ExcelIgnore()
            def secret(self): ...
            def __init__(self, name=None, age=None, secret=None):
                self.name = name; self.age = age; self.secret = secret
        cols = parse_excel_columns(Func)
        assert [(c.attr_name, c.header) for c in cols] == [("name", "姓名"), ("age", "年龄")]

    def test_parse_columns_all_ignored_raises(self):
        class AllIgnored:
            a = ExcelIgnore()
        with pytest.raises(ExcelPropertyError):
            parse_excel_columns(AllIgnored)


# ==================== 转换器 ====================

class TestConverters:
    def test_integer_converter(self):
        c = IntegerConverter()
        assert c.from_excel(12) == 12
        assert c.from_excel("12.0") == 12
        assert c.from_excel("") is None
        assert c.from_excel(None) is None
        assert c.to_excel(3) == 3

    def test_float_converter(self):
        c = FloatConverter()
        assert c.from_excel("3.14") == 3.14
        assert c.to_excel(2) == 2.0

    def test_boolean_converter(self):
        c = BooleanConverter()
        assert c.from_excel(True) is True
        assert c.from_excel("是") is True
        assert c.from_excel("0") is False
        assert c.from_excel("否") is False
        assert c.to_excel(True) is True

    def test_string_converter(self):
        c = StringConverter()
        assert c.from_excel("  x ") == "x"
        assert c.to_excel(None) == ""

    def test_date_string_converter(self):
        c = DateStringConverter("%Y-%m-%d")
        assert c.to_excel(datetime(2026, 8, 9)) == "2026-08-09"
        assert c.from_excel("2026-08-09") == datetime(2026, 8, 9)
        assert c.from_excel("not-a-date") is None

    def test_big_decimal_converter(self):
        c = BigDecimalConverter()
        assert c.to_excel(Decimal("123.45")) == "123.45"
        assert c.from_excel("123.45") == Decimal("123.45")
        assert c.from_excel(None) is None

    def test_resolve_converter_by_type(self):
        assert isinstance(resolve_converter(int), IntegerConverter)
        assert isinstance(resolve_converter(float), FloatConverter)
        assert isinstance(resolve_converter(bool), BooleanConverter)
        assert isinstance(resolve_converter(Decimal), BigDecimalConverter)

    def test_resolve_converter_optional_type(self):
        from typing import Optional
        c = resolve_converter(Optional[int])
        assert isinstance(c, IntegerConverter)

    def test_resolve_converter_declared_overrides(self):
        c = resolve_converter(int, declared=StringConverter)
        assert isinstance(c, StringConverter)

    def test_resolve_converter_date_format_injected(self):
        c = resolve_converter(datetime, date_format="%Y/%m/%d")
        assert isinstance(c, DateStringConverter)
        assert c.fmt == "%Y/%m/%d"

    def test_custom_converter(self, tmp_path):
        import openpyxl

        class UpperConverter(Converter):
            def to_excel(self, value):
                return str(value).upper() if value else value
            def from_excel(self, cell_value):
                return str(cell_value).lower() if cell_value else cell_value

        @ExcelSheet("s")
        class C:
            name = ExcelProperty("名", order=1, converter=UpperConverter)
            def __init__(self, name=None):
                self.name = name

        f = tmp_path / "custom_conv.xlsx"
        write_excel(str(f), C, [C("alice")])

        # 验证 to_excel 生效：写入单元格为大写
        wb = openpyxl.load_workbook(str(f))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "ALICE"
        wb.close()

        # 验证 from_excel 生效：读回为小写
        rows = read_excel(str(f), C)
        assert rows[0].name == "alice"


# ==================== 读写 round-trip ====================

class TestReadWrite:
    def _data(self):
        return [
            DemoUser(76543210987654321, "张三", 28, Decimal("123.45"),
                     datetime(2026, 8, 9, 12, 0, 0), True, "ignore1"),
            DemoUser(2, "李四", 35, Decimal("999.00"),
                     datetime(2026, 1, 2, 3, 4, 5), False, "ignore2"),
        ]

    def test_write_and_read_round_trip(self, tmp_path):
        f = tmp_path / "rt.xlsx"
        EasyExcel.write(str(f), head=DemoUser).sheet("用户列表").doWrite(self._data())
        rows = EasyExcel.read(str(f), head=DemoUser).doRead()
        assert len(rows) == 2
        assert rows[0].name == "张三"
        assert rows[0].age == 28
        assert rows[0].amount == Decimal("123.45")
        assert rows[0].created_at == datetime(2026, 8, 9, 12, 0, 0)
        assert rows[0].active is True
        assert rows[1].active is False

    def test_big_number_preserved_as_string(self, tmp_path):
        f = tmp_path / "bignum.xlsx"
        EasyExcel.write(str(f), head=DemoUser).doWrite(self._data())
        rows = read_excel(str(f), DemoUser)
        # 大数字以字符串写入，保留全部 17 位（Excel 数值会截断到 15 位）
        assert rows[0].id == "76543210987654321"
        assert len(rows[0].id) == 17

    def test_excel_ignore_field_skipped(self, tmp_path):
        f = tmp_path / "ignore.xlsx"
        EasyExcel.write(str(f), head=DemoUser).doWrite(self._data())
        # remark 列不应出现在表头
        import openpyxl
        wb = openpyxl.load_workbook(str(f))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "备注" not in headers and None not in headers
        wb.close()

    def test_header_order(self, tmp_path):
        f = tmp_path / "order.xlsx"
        EasyExcel.write(str(f), head=DemoUser).doWrite(self._data())
        import openpyxl
        wb = openpyxl.load_workbook(str(f))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers == ["ID", "姓名", "年龄", "金额", "创建时间", "是否启用"]
        wb.close()

    def test_fallback_plain_model(self, tmp_path):
        f = tmp_path / "plain.xlsx"
        users = [PlainUser(1, "alice", "a@x.com"), PlainUser(2, "bob", "b@x.com")]
        write_excel(str(f), PlainUser, users)
        rows = read_excel(str(f), PlainUser)
        assert len(rows) == 2
        assert rows[0].username == "alice"
        assert rows[1].email == "b@x.com"

    def test_convenience_read_write(self, tmp_path):
        f = tmp_path / "conv.xlsx"
        write_excel(str(f), DemoUser, self._data())
        rows = read_excel(str(f), DemoUser)
        assert len(rows) == 2
        assert isinstance(rows[0].amount, Decimal)

    def test_empty_rows_skipped(self, tmp_path):
        f = tmp_path / "empty.xlsx"
        EasyExcel.write(str(f), head=DemoUser).doWrite(self._data())
        # 追加一个全空行（手动构造）
        import openpyxl
        wb = openpyxl.load_workbook(str(f))
        ws = wb.active
        ws.append([None, None, None, None, None, None])
        wb.save(str(f)); wb.close()
        rows = read_excel(str(f), DemoUser)
        assert len(rows) == 2  # 空行被跳过


# ==================== 多 sheet ====================

class TestMultipleSheets:
    def test_write_all_multiple_sheets(self, tmp_path):
        f = tmp_path / "multi.xlsx"
        d1 = [DemoUser(1, "a", 1, Decimal("1"), datetime(2026, 1, 1), True, "x")]
        d2 = [DemoUser(2, "b", 2, Decimal("2"), datetime(2026, 1, 2), False, "y")]
        EasyExcel.write(str(f), head=DemoUser).doWriteAll({"S1": d1, "S2": d2})
        result = EasyExcel.read(str(f), head=DemoUser).doReadAll()
        assert set(result.keys()) == {"S1", "S2"}
        assert result["S1"][0].name == "a"
        assert result["S2"][0].name == "b"

    def test_read_sheet_by_name(self, tmp_path):
        f = tmp_path / "byname.xlsx"
        EasyExcel.write(str(f), head=DemoUser).doWriteAll({"Alpha": [], "Beta": [DemoUser(9, "z", 9, Decimal("9"), datetime(2026, 1, 1), True, "")]})
        rows = EasyExcel.read(str(f), head=DemoUser).sheet(sheet_name="Beta").doRead()
        assert len(rows) == 1
        assert rows[0].name == "z"

    def test_read_sheet_by_index(self, tmp_path):
        f = tmp_path / "byidx.xlsx"
        EasyExcel.write(str(f), head=DemoUser).doWriteAll({"First": [DemoUser(1, "a", 1, Decimal("1"), datetime(2026, 1, 1), True, "")], "Second": []})
        rows = EasyExcel.read(str(f), head=DemoUser).sheet(sheet_no=0).doRead()
        assert len(rows) == 1
        assert rows[0].name == "a"

    def test_read_nonexistent_sheet_raises(self, tmp_path):
        f = tmp_path / "nope.xlsx"
        write_excel(str(f), DemoUser, [DemoUser(1, "a", 1, Decimal("1"), datetime(2026, 1, 1), True, "")])
        with pytest.raises(ExcelReadError):
            EasyExcel.read(str(f), head=DemoUser).sheet(sheet_name="不存在").doRead()


# ==================== 配置与降级 ====================

class TestConfigAndDegrade:
    def test_head_row_number(self, tmp_path):
        """表头不在第 1 行：前置说明行 + 表头在第 2 行。"""
        import openpyxl
        f = tmp_path / "head2.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws.append(["这是一份说明，不是表头", None, None])  # row1 占位
        ws.append(["姓名", "年龄", "金额"])                 # row2 表头
        ws.append(["ann", "30", "1.5"])
        wb.save(str(f)); wb.close()

        @ExcelSheet(head_row_number=2)
        class C:
            name = ExcelProperty("姓名", order=1)
            age = ExcelProperty("年龄", order=2)
            amount = ExcelProperty("金额", order=3, converter=BigDecimalConverter)
            def __init__(self, name=None, age=None, amount=None):
                self.name = name; self.age = age; self.amount = amount
        rows = EasyExcel.read(str(f), head=C).doRead()
        assert rows[0].name == "ann"
        assert rows[0].amount == Decimal("1.5")

    def test_fluent_builder_returns_self(self):
        r = EasyExcel.read("x", head=DemoUser)
        assert r.head_row_number(2) is r
        assert r.sheet(sheet_no=0) is r
        w = EasyExcel.write("x", head=DemoUser)
        assert w.sheet("S") is w

    def test_write_without_head_raises(self, tmp_path):
        f = tmp_path / "nohead.xlsx"
        with pytest.raises(ExcelWriteError):
            EasyExcel.write(str(f)).doWrite([1, 2])

    def test_read_without_head_raises(self, tmp_path):
        f = tmp_path / "rh.xlsx"
        write_excel(str(f), DemoUser, [DemoUser(1, "a", 1, Decimal("1"), datetime(2026, 1, 1), True, "")])
        with pytest.raises(ExcelReadError):
            EasyExcel.read(str(f)).doRead()

    def test_dependency_error_when_openpyxl_missing(self, tmp_path, monkeypatch):
        """模拟 openpyxl 未安装：read/write 抛 ExcelDependencyError。"""
        import builtins
        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("simulated missing openpyxl")
            return real_import(name, *args, **kwargs)
        monkeypatch.setattr(builtins, "__import__", fake_import)

        with pytest.raises(ExcelDependencyError):
            EasyExcel.read(str(tmp_path / "x.xlsx"), head=DemoUser).doRead()
        with pytest.raises(ExcelDependencyError):
            EasyExcel.write(str(tmp_path / "x.xlsx"), head=DemoUser).doWrite([])

    def test_annotation_works_without_openpyxl(self, monkeypatch):
        """注解声明不依赖 openpyxl：模拟未安装仍可声明 @ExcelProperty。"""
        import builtins
        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("simulated missing openpyxl")
            return real_import(name, *args, **kwargs)
        monkeypatch.setattr(builtins, "__import__", fake_import)

        @ExcelSheet("ok")
        class NoDep:
            name = ExcelProperty("姓名", order=1)
            def __init__(self, name=None):
                self.name = name
        cols = parse_excel_columns(NoDep)
        assert cols[0].header == "姓名"


# ==================== 样式与格式 ====================

class TestStyleAndFormat:
    def test_freeze_head_and_style(self, tmp_path):
        f = tmp_path / "style.xlsx"
        EasyExcel.write(str(f), head=DemoUser).doWrite(
            [DemoUser(1, "a", 1, Decimal("1"), datetime(2026, 1, 1), True, "")])
        import openpyxl
        wb = openpyxl.load_workbook(str(f))
        ws = wb.active
        assert ws.freeze_panes is not None  # 冻结表头生效
        head_cell = ws.cell(row=1, column=1)
        assert head_cell.font.bold is True  # 表头加粗
        wb.close()

    def test_num_format_applied(self, tmp_path):
        @ExcelSheet("s")
        class C:
            price = ExcelProperty("价格", order=1, num_format="#,##0.00")
            def __init__(self, price=None):
                self.price = price
        f = tmp_path / "numfmt.xlsx"
        write_excel(str(f), C, [C(1234.5)])
        import openpyxl
        wb = openpyxl.load_workbook(str(f))
        ws = wb.active
        cell = ws.cell(row=2, column=1)
        assert cell.number_format == "#,##0.00"
        assert cell.value == 1234.5
        wb.close()

    def test_custom_column_width(self, tmp_path):
        f = tmp_path / "width.xlsx"
        EasyExcel.write(str(f), head=DemoUser).doWrite(
            [DemoUser(1, "a", 1, Decimal("1"), datetime(2026, 1, 1), True, "")])
        import openpyxl
        wb = openpyxl.load_workbook(str(f))
        ws = wb.active
        # name 列设置了 width=12
        col_letter = ws.cell(row=1, column=2).column_letter
        assert ws.column_dimensions[col_letter].width == 12.0
        wb.close()


# ==================== ORM 风格类型注解（对齐 @entity） ====================

class TestOrmStyleAnnotations:
    """ORM 风格：类型注解字段自动建列 + @excel_sheet 自动生成 __init__。"""

    def test_auto_init_generated(self):
        """@excel_sheet 自动生成 __init__，无需手写。"""
        @ExcelSheet("测试")
        class Demo:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""
            age: int = 0

        obj = Demo(id=1, name="张三", age=28)
        assert obj.id == 1
        assert obj.name == "张三"
        assert obj.age == 28

    def test_auto_init_default_values(self):
        """未传参时使用类型注解声明的默认值。"""
        @ExcelSheet("测试")
        class Demo:
            name: str = "默认名"
            age: int = 18
            score: float = 0.0

        obj = Demo()
        assert obj.name == "默认名"
        assert obj.age == 18
        assert obj.score == 0.0

    def test_auto_init_rejects_unknown_fields(self):
        """自动生成的 __init__ 拒绝未知字段。"""
        @ExcelSheet("测试")
        class Demo:
            name: str = ""

        with pytest.raises(TypeError, match="Unexpected field"):
            Demo(unknown_field=1)

    def test_existing_init_preserved(self):
        """类已有 __init__ 时不被覆盖。"""
        @ExcelSheet("测试")
        class Demo:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""

            def __init__(self, custom_id=None):
                self.id = custom_id
                self.name = "固定值"

        obj = Demo(custom_id=99)
        assert obj.id == 99
        assert obj.name == "固定值"
        # 不支持关键字 id=（因为 __init__ 参数是 custom_id）
        with pytest.raises(TypeError):
            Demo(id=1)

    def test_auto_columns_from_annotations(self):
        """类型注解字段自动建列（无 ExcelProperty 也自动映射）。"""
        @ExcelSheet("测试")
        class Demo:
            id: int = ExcelProperty("工号", order=1)
            name: str = ""          # 自动建列
            age: int = 0            # 自动建列
            salary: float = 0.0     # 自动建列

        cols = parse_excel_columns(Demo)
        attr_names = [c.attr_name for c in cols]
        assert attr_names == ["id", "name", "age", "salary"]

        # id 有显式表头，其余按字段名生成
        headers = [c.header for c in cols]
        assert headers == ["工号", "Name", "Age", "Salary"]

    def test_type_inference_from_annotations(self):
        """类级类型注解用于转换器自动选择（无 __init__ 时也能推断类型）。"""
        @ExcelSheet("测试")
        class Demo:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""
            age: int = 0
            score: float = 0.0
            active: bool = True

        cols = parse_excel_columns(Demo)
        col_map = {c.attr_name: c for c in cols}
        assert col_map["id"].py_type is int
        assert col_map["name"].py_type is str
        assert col_map["age"].py_type is int
        assert col_map["score"].py_type is float
        assert col_map["active"].py_type is bool

    def test_ignore_field_in_orm_style(self):
        """ORM 风格类中 ExcelIgnore 字段被跳过。"""
        @ExcelSheet("测试")
        class Demo:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""
            remark = ExcelIgnore()      # 跳过

        cols = parse_excel_columns(Demo)
        attr_names = [c.attr_name for c in cols]
        assert "remark" not in attr_names
        assert attr_names == ["id", "name"]

    def test_private_field_skipped(self):
        """以 _ 开头的私有字段不参与导出。"""
        @ExcelSheet("测试")
        class Demo:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""
            _cache: dict = {}           # 私有，跳过

        cols = parse_excel_columns(Demo)
        assert all(not c.attr_name.startswith("_") for c in cols)
        assert {c.attr_name for c in cols} == {"id", "name"}

    def test_mixed_explicit_and_auto_order(self):
        """显式 ExcelProperty 有 order，自动列按声明顺序排列。"""
        @ExcelSheet("测试")
        class Demo:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""          # 自动列，order 回退为声明顺序
            age: int = 0

        cols = parse_excel_columns(Demo)
        # id: order=1(显式), name: order=decl_order=1(回退), age: order=decl_order=2(回退)
        # 排序：index(None→inf) → order → decl_order
        # id(1,decl=0) < name(1,decl=1) < age(2,decl=2)
        assert [c.attr_name for c in cols] == ["id", "name", "age"]

    def test_round_trip_orm_style(self, tmp_path):
        """ORM 风格类的写入 + 读取 round-trip。"""
        @ExcelSheet("员工列表")
        class Employee:
            id: int = ExcelProperty("工号", order=1, big_number=True)
            name: str = ""
            age: int = 0
            salary: float = 0.0

        data = [
            Employee(id=1, name="张三", age=28, salary=9999.5),
            Employee(id=2, name="李四", age=35, salary=12345.0),
        ]
        f = tmp_path / "orm_rt.xlsx"
        EasyExcel.write(str(f), head=Employee).sheet("员工列表").doWrite(data)
        rows = EasyExcel.read(str(f), head=Employee).doRead()

        assert len(rows) == 2
        assert rows[0].name == "张三"
        assert rows[0].age == 28
        assert rows[0].salary == 9999.5
        assert rows[1].name == "李四"
        assert rows[1].salary == 12345.0

    def test_round_trip_with_ignore(self, tmp_path):
        """ORM 风格 + ExcelIgnore 的 round-trip（忽略字段不导出）。"""
        @ExcelSheet("员工")
        class Employee:
            id: int = ExcelProperty("工号", order=1)
            name: str = ""
            internal_note = ExcelIgnore()

        data = [
            Employee(id=1, name="张三", internal_note="机密"),
            Employee(id=2, name="李四", internal_note="秘密"),
        ]
        f = tmp_path / "orm_ignore.xlsx"
        write_excel(str(f), Employee, data)
        rows = read_excel(str(f), Employee)

        assert len(rows) == 2
        assert rows[0].id == 1
        assert rows[0].name == "张三"
        # internal_note 不在导出列中，读回为默认值 None
        assert rows[0].internal_note is None

    def test_no_init_no_decorator_still_works(self, tmp_path):
        """无 @excel_sheet、无 __init__ 的纯注解类也能解析（reader 用 object.__new__ 回退）。"""
        class Demo:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""

        cols = parse_excel_columns(Demo)
        assert {c.attr_name for c in cols} == {"id", "name"}
        assert cols[0].py_type is int
        assert cols[1].py_type is str

    def test_inheritance_orm_style(self):
        """ORM 风格支持继承（子类注解 + 父类注解合并）。"""
        @ExcelSheet("基类")
        class Base:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""

        @ExcelSheet("子类")
        class Child(Base):
            age: int = 0
            extra: str = "默认"

        cols = parse_excel_columns(Child)
        attr_names = [c.attr_name for c in cols]
        assert "id" in attr_names
        assert "name" in attr_names
        assert "age" in attr_names
        assert "extra" in attr_names

    def test_backward_compat_excel_sheet_alias(self):
        """旧名 @excel_sheet 仍可用（向后兼容别名）。"""
        @excel_sheet("兼容测试")
        class Demo:
            id: int = ExcelProperty("ID", order=1)
            name: str = ""

        obj = Demo(id=1, name="x")
        assert obj.id == 1
        assert obj.name == "x"
        meta = Demo.__excel_sheet__
        assert isinstance(meta, ExcelSheet)
        assert meta.sheet_name == "兼容测试"


# ==================== 组合式注解（@Entity + @ExcelSheet + @CsvFile） ====================

class TestCompositionalAnnotations:
    """组合式：同一类上同时使用 ORM + Excel + CSV 注解。"""

    def test_descriptors_not_replaced_by_orm(self):
        """ORM _auto_infer_columns 不覆盖 ExcelProperty/CsvProperty 描述符。"""
        from spring.orm import Entity, Id, Column

        @Entity("users")
        @ExcelSheet("用户列表")
        class Demo:
            id: int = Id()
            name: str = Column("name", default="")
            email: str = ExcelProperty("邮箱")

        # ExcelProperty 描述符未被替换为 Column
        assert isinstance(Demo.__dict__["email"], ExcelProperty)
        assert Demo.__dict__["email"].value == "邮箱"
        # Column 描述符保留
        assert isinstance(Demo.__dict__["name"], Column)
        # Id 描述符保留
        assert type(Demo.__dict__["id"]).__name__ == "Id"

    def test_auto_init_with_foreign_descriptors(self):
        """组合式自动 __init__ 正确处理跨模块描述符的默认值。"""
        from spring.orm import Entity, Id, Column
        from spring.csv import CsvFile, CsvProperty

        @Entity("users")
        @ExcelSheet("用户列表")
        @CsvFile("users.csv")
        class User:
            id: int = Id()
            name: str = Column("name", default="默认名")
            age: int = 0
            email: str = ExcelProperty("邮箱")
            phone: str = CsvProperty("手机")

        u = User()
        assert u.id is None        # Id 的 default 属性为 None
        assert u.name == "默认名"   # Column 的 default 属性
        assert u.age == 0          # 普通默认值
        assert u.email is None     # ExcelProperty 无 default → None
        assert u.phone is None     # CsvProperty 无 default → None

    def test_all_fields_in_excel(self):
        """组合式类所有字段都出现在 Excel 列中。"""
        from spring.orm import Entity, Id, Column
        from spring.csv import CsvFile, CsvProperty

        @Entity("users")
        @ExcelSheet("用户列表")
        @CsvFile("users.csv")
        class User:
            id: int = Id()
            name: str = Column("name")
            age: int = 0
            email: str = ExcelProperty("邮箱")
            phone: str = CsvProperty("手机")

        cols = parse_excel_columns(User)
        attr_names = {c.attr_name for c in cols}
        assert attr_names == {"id", "name", "age", "email", "phone"}

        # email 保留 ExcelProperty 的表头
        col_map = {c.attr_name: c for c in cols}
        assert col_map["email"].header == "邮箱"
        # phone 自动建列，表头按字段名
        assert col_map["phone"].header == "Phone"

    def test_excel_round_trip_compositional(self, tmp_path):
        """组合式类的 Excel 写入 + 读取 round-trip。"""
        from spring.orm import Entity, Id, Column
        from spring.csv import CsvFile, CsvProperty

        @Entity("users")
        @ExcelSheet("用户列表")
        @CsvFile("users.csv")
        class User:
            id: int = Id()
            name: str = Column("name", default="")
            age: int = 0
            email: str = ExcelProperty("邮箱")
            phone: str = CsvProperty("手机")

        data = [
            User(id=1, name="张三", age=28, email="a@b.com", phone="138"),
            User(id=2, name="李四", age=35, email="c@d.com", phone="139"),
        ]
        f = tmp_path / "comp.xlsx"
        EasyExcel.write(str(f), head=User).sheet("用户列表").doWrite(data)
        rows = EasyExcel.read(str(f), head=User).doRead()

        assert len(rows) == 2
        assert rows[0].id == 1
        assert rows[0].name == "张三"
        assert rows[0].age == 28
        assert rows[0].email == "a@b.com"
        assert rows[0].phone == "138"
        assert rows[1].name == "李四"
        assert rows[1].email == "c@d.com"
