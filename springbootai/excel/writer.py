"""SpringBootAI Excel 写入引擎。

按 ``@ExcelProperty`` / ``@ExcelIgnore`` / ``@excel_sheet`` 注解把实体对象列表写入 Excel。
底层依赖 openpyxl（仅 ``doWrite`` 时检测，注解声明无需安装）。

写入流程对齐 alibaba EasyExcel：
    1. 按列模型排序写表头行（默认带样式）。
    2. 逐行按 ``converter.to_excel`` 转换并写单元格；应用数字格式、大数字防丢精度。
    3. 冻结表头、自适应列宽（可配置）。
"""
from __future__ import annotations

from typing import Any, Dict, Iterable, List, Optional, Type

from .annotations import (
    ExcelColumnModel, ExcelSheet, _get_class_sheet_meta, parse_excel_columns,
)
from .converters import resolve_converter
from .exceptions import ExcelDependencyError, ExcelWriteError

# Excel 有效数字位数上限为 15 位，超过即丢精度；长 ID/大数字按字符串写入
_EXCEL_MAX_SIGNIFICANT_DIGITS = 15


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError as e:
        raise ExcelDependencyError(
            "Excel 写入依赖 openpyxl，请先安装：pip install springbootAI[excel]"
        ) from e


def _get_value(item: Any, attr_name: str) -> Any:
    """从实体对象或 dict 取字段值。"""
    if isinstance(item, dict):
        return item.get(attr_name)
    return getattr(item, attr_name, None)


def _is_large_int(value: Any) -> bool:
    """是否为会丢精度的长整数（>15 位有效数字）。"""
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return len(str(abs(value))) > _EXCEL_MAX_SIGNIFICANT_DIGITS
    return False


class ExcelWriter:
    """Excel 写入器。由 ``EasyExcel.write(...)`` 构建，调用 ``doWrite`` 执行。"""

    def __init__(
        self,
        target: Any,
        head: Optional[Type] = None,
        sheet_name: Optional[str] = None,
    ):
        self.target = target
        self.head = head
        self._sheet_name = sheet_name
        self._in_memory: bool = False

    # ---- 流式配置 ----

    def sheet(self, sheet_name: str = "Sheet1") -> "ExcelWriter":
        self._sheet_name = sheet_name
        return self

    def in_memory(self, flag: bool = True) -> "ExcelWriter":
        """写入内存 Workbook（用于随后取 workbook 对象，而非落盘）。"""
        self._in_memory = flag
        return self

    # ---- 执行 ----

    def doWrite(self, data: Iterable[Any]) -> Any:
        """写入单个工作表并保存到 ``target``。返回目标路径或内存 workbook。"""
        openpyxl = _require_openpyxl()
        if self.head is None:
            raise ExcelWriteError("write 需指定 head 实体类")
        wb = openpyxl.Workbook()
        # 移除默认空 Sheet
        default_ws = wb.active
        sheet_name = self._sheet_name or _get_class_sheet_meta(self.head).sheet_name or "Sheet1"
        ws = wb.create_sheet(title=sheet_name)
        wb.remove(default_ws)
        self._write_sheet(ws, self.head, list(data))
        return self._save(wb)

    def doWriteAll(self, sheets: Dict[str, Iterable[Any]], head: Optional[Type] = None) -> Any:
        """写入多个工作表。``sheets`` 为 {sheet_name: data_list}，所有 sheet 共用同一 head。"""
        openpyxl = _require_openpyxl()
        head_cls = head or self.head
        if head_cls is None:
            raise ExcelWriteError("doWriteAll 需指定 head 实体类")
        wb = openpyxl.Workbook()
        default_ws = wb.active
        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            self._write_sheet(ws, head_cls, list(data))
        wb.remove(default_ws)
        return self._save(wb)

    # ---- 内部 ----

    def _write_sheet(self, ws, head_cls: Type, data: List[Any]) -> None:
        from .style import apply_content_style, apply_head_style

        sheet_meta = _get_class_sheet_meta(head_cls)
        columns: List[ExcelColumnModel] = parse_excel_columns(head_cls)
        head_row = sheet_meta.head_row_number

        # 1. 表头
        for offset, model in enumerate(columns):
            col_idx = offset + 1
            cell = ws.cell(row=head_row, column=col_idx, value=model.header)
            apply_head_style(cell, model.head_style or sheet_meta.head_style or "default")

        # 2. 数据行
        max_lengths = {m.attr_name: len(str(m.header)) for m in columns}
        for row_offset, item in enumerate(data, start=1):
            row_idx = head_row + row_offset
            for offset, model in enumerate(columns):
                col_idx = offset + 1
                raw = _get_value(item, model.attr_name)
                cell_value, as_string = self._convert_to_cell(model, raw)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                apply_content_style(cell, model.content_style or sheet_meta.content_style or "default")
                # 数字格式（仅对数值单元格生效）
                if model.num_format and not as_string and isinstance(cell_value, (int, float)):
                    cell.number_format = model.num_format
                # 统计列宽
                display_len = len(str(cell_value)) if cell_value is not None else 0
                if display_len > max_lengths[model.attr_name]:
                    max_lengths[model.attr_name] = display_len

        # 3. 冻结表头
        if sheet_meta.freeze_head:
            ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

        # 4. 列宽
        self._apply_column_width(ws, columns, max_lengths, sheet_meta)

    def _convert_to_cell(self, model: ExcelColumnModel, raw: Any):
        """Python 值 -> (单元格值, 是否强制字符串)。"""
        converter = resolve_converter(model.py_type, model.converter, model.date_format)
        if converter is not None:
            try:
                converted = converter.to_excel(raw)
            except Exception as e:
                raise ExcelWriteError(
                    f"字段 '{model.attr_name}' 转换失败 (值={raw!r}): {e}"
                ) from e
            as_string = isinstance(converted, str)
            return converted, as_string
        # 无转换器：按规则处理
        if raw is None:
            return None, False
        if model.big_number or _is_large_int(raw):
            return str(raw), True
        return raw, False

    def _apply_column_width(self, ws, columns: List[ExcelColumnModel],
                            max_lengths: dict, sheet_meta: ExcelSheet) -> None:
        for offset, model in enumerate(columns):
            col_idx = offset + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            if model.width and model.width > 0:
                ws.column_dimensions[col_letter].width = float(model.width)
            elif sheet_meta.auto_width:
                # 自适应：max(表头/内容长度) + 2 余量，上限 60
                ws.column_dimensions[col_letter].width = min(max(max_lengths[model.attr_name] + 2, 10), 60)

    def _save(self, wb) -> Any:
        target = self.target
        if self._in_memory:
            return wb
        try:
            if isinstance(target, (str, bytes)) or hasattr(target, "__fspath__"):
                wb.save(filename=str(target))
                return target
            # 类文件对象
            wb.save(target)
            return target
        except Exception as e:
            raise ExcelWriteError(f"保存 Excel 文件失败: {e}") from e
        finally:
            try:
                if not self._in_memory:
                    wb.close()
            except Exception:
                pass


__all__ = ["ExcelWriter"]
