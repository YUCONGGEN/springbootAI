"""SpringBootAI Excel 默认样式 —— 轻量表头/内容样式（基于 openpyxl）。

仅在 writer 实际写入时使用，且为可选项。样式名通过 ``@ExcelSheet(head_style=...)`` 或
``@ExcelProperty(head_style=...)`` 指定；未指定时使用 ``DEFAULT_HEAD``。

设计克制：不提供复杂主题系统，仅给出一个加粗居中、带边框和浅色填充的默认表头样式，
以及一个带边框的默认内容样式，满足"功能齐全"的同时避免过度设计。
"""
from __future__ import annotations

from typing import Any, Dict

# 样式缓存：同名样式复用同一 openpyxl 对象（openpyxl 要求相同样式复用以免撑大文件）
_STYLE_CACHE: Dict[str, Dict[str, Any]] = {}


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return Font, PatternFill, Alignment, Border, Side
    except ImportError as e:  # pragma: no cover - 依赖检测由上层统一处理
        raise e


def get_head_style(name: str = "default"):
    """获取表头样式对象（带缓存）。"""
    from .exceptions import ExcelDependencyError
    try:
        Font, PatternFill, Alignment, Border, Side = _require_openpyxl()
    except ImportError:
        raise ExcelDependencyError(
            "Excel 写入依赖 openpyxl，请先安装：pip install springbootAI[excel]"
        )
    key = f"head::{name}"
    if key in _STYLE_CACHE:
        return _STYLE_CACHE[key]
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    style = {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="4472C4"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": border,
    }
    _STYLE_CACHE[key] = style
    return style


def get_content_style(name: str = "default"):
    """获取内容样式对象（带缓存）。"""
    from .exceptions import ExcelDependencyError
    try:
        Font, PatternFill, Alignment, Border, Side = _require_openpyxl()
    except ImportError:
        raise ExcelDependencyError(
            "Excel 写入依赖 openpyxl，请先安装：pip install springbootAI[excel]"
        )
    key = f"content::{name}"
    if key in _STYLE_CACHE:
        return _STYLE_CACHE[key]
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    style = {
        "font": Font(size=11),
        "alignment": Alignment(vertical="center", wrap_text=False),
        "border": border,
    }
    _STYLE_CACHE[key] = style
    return style


def apply_head_style(cell, name: str = "default") -> None:
    """把表头样式应用到单元格。"""
    style = get_head_style(name)
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]
    cell.border = style["border"]


def apply_content_style(cell, name: str = "default") -> None:
    """把内容样式应用到单元格。"""
    style = get_content_style(name)
    cell.font = style["font"]
    cell.alignment = style["alignment"]
    cell.border = style["border"]


__all__ = [
    "get_head_style",
    "get_content_style",
    "apply_head_style",
    "apply_content_style",
]
