"""SpringPy Excel 读取引擎。

按 ``@ExcelProperty`` / ``@ExcelIgnore`` / ``@excel_sheet`` 注解把 Excel 工作表读为实体对象列表。
底层依赖 openpyxl（仅 ``doRead`` 时检测，注解声明无需安装）。

读取流程对齐 alibaba EasyExcel：
    1. 读取表头行（``head_row_number``，默认 1）。
    2. 表头文案 -> ``ExcelColumnModel`` 映射（按 ``value`` 匹配；无注解时按列位置匹配）。
    3. 数据行逐行按 ``converter.from_excel`` 转换，构造实体实例。
"""
from __future__ import annotations

import inspect
from typing import Any, List, Optional, Type, Union

from .annotations import (
    ExcelColumnModel, ExcelSheet, _get_class_sheet_meta, parse_excel_columns,
)
from .converters import resolve_converter
from .exceptions import ExcelDependencyError, ExcelReadError


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError as e:
        raise ExcelDependencyError(
            "Excel 读取依赖 openpyxl，请先安装：pip install springpy[excel]"
        ) from e


def _build_instance(cls: Type, kwargs: dict) -> Any:
    """构造实体实例：优先 ``cls(**kwargs)``，失败则逐字段 setattr（兼容纯 ``__init__`` 模型）。"""
    # 只保留 __init__ 能接收的参数
    try:
        sig = inspect.signature(cls.__init__)
        params = list(sig.parameters.values())[1:]  # 跳过 self
        accept_kwargs = any(p.kind == inspect.Parameter.VAR_KEYWORD for p in params)
        if accept_kwargs:
            return cls(**kwargs)
        allowed = {p.name for p in params if p.kind in (
            inspect.Parameter.POSITIONAL_OR_KEYWORD, inspect.Parameter.KEYWORD_ONLY,
        )}
        filtered = {k: v for k, v in kwargs.items() if k in allowed}
        return cls(**filtered)
    except TypeError:
        pass
    # 回退：无参构造 + setattr
    try:
        obj = cls()
    except Exception:
        obj = object.__new__(cls)
    for k, v in kwargs.items():
        try:
            setattr(obj, k, v)
        except Exception:
            pass
    return obj


class ExcelReader:
    """Excel 读取器。由 ``EasyExcel.read(...)`` 构建，调用 ``doRead`` 执行。"""

    def __init__(
        self,
        source: Any,
        head: Optional[Type] = None,
        head_row_number: Optional[int] = None,
        sheet_no: Optional[int] = None,
        sheet_name: Optional[str] = None,
    ):
        self.source = source
        self.head = head
        self._head_row_number = head_row_number
        self._sheet_no = sheet_no
        self._sheet_name = sheet_name

    # ---- 流式配置 ----

    def head_row_number(self, n: int) -> "ExcelReader":
        self._head_row_number = n
        return self

    def sheet(self, sheet_no: Optional[int] = None, sheet_name: Optional[str] = None) -> "ExcelReader":
        self._sheet_no = sheet_no
        self._sheet_name = sheet_name
        return self

    # ---- 执行 ----

    def _resolve_sheet(self, wb) -> List:
        """返回要读取的工作表列表。"""
        if self._sheet_name is not None:
            if self._sheet_name not in wb.sheetnames:
                raise ExcelReadError(f"工作表不存在: {self._sheet_name}")
            return [wb[self._sheet_name]]
        if self._sheet_no is not None:
            if self._sheet_no >= len(wb.sheetnames):
                raise ExcelReadError(f"工作表索引越界: {self._sheet_no}")
            return [wb.worksheets[self._sheet_no]]
        # 默认读所有 sheet
        return list(wb.worksheets)

    def _resolve_head_row_number(self, sheet_meta: ExcelSheet) -> int:
        if self._head_row_number is not None:
            return self._head_row_number
        return sheet_meta.head_row_number

    def _read_one_sheet(self, ws) -> List[Any]:
        head_cls = self.head
        if head_cls is None:
            raise ExcelReadError("read 需指定 head 实体类")

        sheet_meta = _get_class_sheet_meta(head_cls)
        columns: List[ExcelColumnModel] = parse_excel_columns(head_cls)
        head_row = self._resolve_head_row_number(sheet_meta)

        # 读取表头行
        max_col = ws.max_column or 0
        max_row = ws.max_row or 0
        if max_row < head_row:
            return []
        header_cells = [ws.cell(row=head_row, column=c).value for c in range(1, max_col + 1)]

        # 列映射：列序号(1-based) -> ExcelColumnModel
        col_mapping = self._map_columns(columns, header_cells)

        # 判断是否走"按位置回退"（类无任何 ExcelProperty 注解）
        used_positional = not _has_explicit_properties(head_cls)

        results: List[Any] = []
        for row_idx in range(head_row + 1, max_row + 1):
            # 跳过全空行
            row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
            if all(v is None or v == "" for v in row_values):
                continue
            kwargs = {}
            for col_idx, model in col_mapping.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                value = self._convert_from_cell(model, cell_value)
                kwargs[model.attr_name] = value
            # 位置回退：未映射的列按声明顺序补齐（仅当类无注解且字段数匹配时）
            if used_positional:
                # 已按位置映射，无需额外处理
                pass
            results.append(_build_instance(head_cls, kwargs))
        return results

    def _map_columns(self, columns: List[ExcelColumnModel], header_cells: List) -> dict:
        """表头 -> 列模型映射。返回 {列序号(1-based): ExcelColumnModel}。"""
        col_mapping: dict = {}
        explicit = any(c.header for c in columns) and _has_explicit_properties(self.head)

        if explicit:
            # 按表头文案匹配
            header_to_col = {}
            for idx, h in enumerate(header_cells, start=1):
                if h is None:
                    continue
                header_to_col.setdefault(str(h).strip(), idx)
            for model in columns:
                col_idx = header_to_col.get(model.header)
                if col_idx is None:
                    # 找不到对应表头，按声明顺序回退到下一个空位
                    continue
                col_mapping[col_idx] = model
        else:
            # 按列位置匹配（无注解或表头为空）
            for offset, model in enumerate(columns):
                col_idx = (model.index if model.index is not None else offset) + 1
                col_mapping[col_idx] = model
        return col_mapping

    def _convert_from_cell(self, model: ExcelColumnModel, cell_value: Any) -> Any:
        converter = resolve_converter(model.py_type, model.converter, model.date_format)
        if converter is None:
            return cell_value
        try:
            return converter.from_excel(cell_value)
        except Exception as e:
            raise ExcelReadError(
                f"字段 '{model.attr_name}' 转换失败 (单元格值={cell_value!r}): {e}"
            ) from e

    def doRead(self) -> List[Any]:
        """读取选定的工作表，返回实体列表。"""
        openpyxl = _require_openpyxl()
        wb = _load_workbook(self.source, openpyxl, read_only=False)
        try:
            sheets = self._resolve_sheet(wb)
            if not sheets:
                return []
            return self._read_one_sheet(sheets[0])
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def doReadAll(self) -> dict:
        """读取所有工作表，返回 {sheet_name: [实体列表]}。"""
        openpyxl = _require_openpyxl()
        wb = _load_workbook(self.source, openpyxl, read_only=False)
        try:
            result = {}
            for ws in wb.worksheets:
                self._sheet_name = ws.title
                result[ws.title] = self._read_one_sheet(ws)
            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass


def _has_explicit_properties(cls: Type) -> bool:
    """类上是否声明了至少一个 ExcelProperty（用于决定按表头还是按位置匹配）。"""
    for base in cls.__mro__:
        for value in vars(base).values():
            if isinstance(value, ExcelColumnModel):
                continue
            if isinstance(value, type):  # 跳过类属性中的类型
                continue
            if hasattr(value, "__excel_property__"):
                return True
            # ExcelProperty 实例作为类属性
    # 再扫一次 ExcelProperty 实例
    for base in cls.__mro__:
        for value in vars(base).values():
            if _is_excel_property(value):
                return True
    return False


def _is_excel_property(value: Any) -> bool:
    from .annotations import ExcelProperty
    # 避免与 ExcelColumnModel 混淆
    return isinstance(value, ExcelProperty)


def _load_workbook(source: Any, openpyxl_module, read_only: bool = False):
    """支持文件路径或类文件对象。"""
    try:
        if isinstance(source, (str, bytes)) or hasattr(source, "__fspath__"):
            return openpyxl_module.load_workbook(filename=str(source), data_only=True, read_only=read_only)
        # 类文件对象
        return openpyxl_module.load_workbook(source, data_only=True, read_only=read_only)
    except Exception as e:
        raise ExcelReadError(f"加载 Excel 文件失败: {e}") from e


__all__ = ["ExcelReader"]
